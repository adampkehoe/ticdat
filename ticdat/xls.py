"""
Read/write ticDat objects from xls files. Requires the xlrd/xlrt module.
PEP8
"""
import ticdat.utils as utils
from ticdat.utils import freezable_factory, TicDatError, verify, containerish, case_space_to_pretty, FrozenDict
import os
from collections import defaultdict
from itertools import product
from ticdat.pandatfactory import PanDatFactory
import datetime
try:
    import xlrd
except:
    xlrd=None
try:
    import xlwt
except:
    xlwt=None
try:
    import xlsxwriter as xlsx
except:
    xlsx=None
try:
    import openpyxl
except:
    openpyxl=None

_can_unit_test = xlrd and xlwt and xlsx

# https://github.com/jmcnamara/XlsxWriter/issues/150 ...
# the xlsxwriter doesn't handle infinity as seamlessly as xls
_longest_sheet = 30

class XlsTicFactory(freezable_factory(object, "_isFrozen")) :
    """
    Primary class for reading/writing Excel files with TicDat objects.
    Your system will need the xlrd package to read .xls and the openpyxl
    to read .xlsx files, the xlwt package to write .xls files, and the
    xlsxwriter package to write .xlsx files.
    Don't create this object explicitly. A XlsTicDatFactory will
    automatically be associated with the xls attribute of the parent
    TicDatFactory.
    """
    def __init__(self, tic_dat_factory):
        """
        Don't create this object explicitly. A XlsTicDatFactory will
        automatically be associated with the xls attribute of the parent
        TicDatFactory.
        :param tic_dat_factory:
        :return:
        """
        self.tic_dat_factory = tic_dat_factory
        self._dv_dt = {}
        self._isFrozen = True
    def create_tic_dat(self, xls_file_path, row_offsets=None, headers_present = True,
                       treat_inf_as_infinity = True,
                       freeze_it = False):
        """
        Create a TicDat object from an Excel file

        :param xls_file_path: An Excel file containing sheets whose names match
                              the table names in the schema.

        :param row_offsets: (optional) A mapping from table names to initial
                            number of rows to skip

        :param headers_present: Boolean. Does the first row of data contain the
                                column headers?

        :param treat_inf_as_infinity: Boolean. Treat the "inf" string (case insensitive) as
                                               as infinity. Similar for "-inf"

        :param freeze_it: boolean. should the returned object be frozen?

        :return: a TicDat object populated by the matching sheets.

        caveats: Missing sheets resolve to an empty table, but missing fields
                 on matching sheets throw an Exception.
                 Sheet names are considered case insensitive, and white space is replaced
                 with underscore for table name matching.
                 Field names are considered case insensitive, but white space is respected.
                 (ticdat supports whitespace in field names but not table names).
                 The following two caveats apply only if data_types are used.
                 --> Any field for which an empty string is invalid data and None is
                     valid data will replace the empty string with None.
                 --> Any field for which must_be_int is true will replace numeric
                     data that satisfies int(x)==x with int(x). In other words,
                     the ticdat equivalent of pandas.read_excel convert_float
                     is to set must_be_int to true in data_types.
        """
        self._verify_differentiable_sheet_names()
        if xls_file_path.endswith('.xlsx'):
            verify(openpyxl, "openpyxl needs to be installed to use this subroutine")
        else:
            verify(xlrd, "xlrd needs to be installed to use this subroutine")
        tdf = self.tic_dat_factory
        verify(not(treat_inf_as_infinity and tdf.generator_tables),
               "treat_inf_as_infinity not implemented for generator tables")
        verify(headers_present or not tdf.generic_tables,
               "headers need to be present to read generic tables")
        verify(utils.DataFrame or not tdf.generic_tables,
               "Strange absence of pandas despite presence of generic tables")
        if self.tic_dat_factory.generic_tables:
            verify(headers_present and treat_inf_as_infinity and not row_offsets,
                   "headers_present, treat_inf_as_infinity and row_offsets must all be at default values\n" +
                   "to use generic tables")
        rtn = self._create_tic_dat_dict(xls_file_path, row_offsets or {}, headers_present, treat_inf_as_infinity)
        if self.tic_dat_factory.generic_tables:
            if xls_file_path.endswith(".xls"):
                print("** Warning : pandas doesn't always play well with older Excel formats.")
            pdf = PanDatFactory(**{t: '*' for t in self.tic_dat_factory.generic_tables})
            pandat = pdf.xls.create_pan_dat(xls_file_path)
            for t in self.tic_dat_factory.generic_tables:
                rtn[t] = getattr(pandat, t)
        rtn = tdf._parameter_table_post_read_adjustment(tdf.TicDat(**rtn))
        if freeze_it:
            return self.tic_dat_factory.freeze_me(rtn)
        return rtn
    def _verify_differentiable_sheet_names(self):
        rtn = defaultdict(set)
        for t in self.tic_dat_factory.all_tables:
            rtn[t[:_longest_sheet]].add(t)
        rtn = [v for k,v in rtn.items() if len(v) > 1]
        verify(not rtn, "The following tables collide when names are truncated to %s characters.\n%s"%
               (_longest_sheet, sorted(map(sorted, rtn))))
    def _get_sheets_and_fields(self, xls_file_path, all_tables, row_offsets, headers_present,
                               print_missing_tables = False):
        verify(utils.stringish(xls_file_path) and os.path.exists(xls_file_path),
               "xls_file_path argument %s is not a valid file path."%xls_file_path)
        if xls_file_path.endswith('.xlsx'):
            try:
                book = openpyxl.load_workbook(xls_file_path)
                book.datemode = 0
            except Exception as e:
                raise TicDatError("Unable to open %s as xlsx file : %s"%(xls_file_path, e))
        else:
            try :
                book = xlrd.open_workbook(xls_file_path)
            except Exception as e:
                raise TicDatError("Unable to open %s as xls file : %s"%(xls_file_path, e))
        sheets = defaultdict(list)
        if xls_file_path.endswith('.xlsx'):
            for table, sheet in product(all_tables, book._sheets):
                if table.lower()[:_longest_sheet] == sheet.title.lower().replace(' ', '_')[:_longest_sheet]:
                    sheets[table].append(sheet)
        else:
            for table, sheet in product(all_tables, book.sheets()) :
                if table.lower()[:_longest_sheet] == sheet.name.lower().replace(' ', '_')[:_longest_sheet]:
                    sheets[table].append(sheet)
        duplicated_sheets = tuple(_t for _t,_s in sheets.items() if len(_s) > 1)
        verify(not duplicated_sheets, "The following sheet names were duplicated : " +
               ",".join(duplicated_sheets))
        if xls_file_path.endswith('.xlsx'):
            sheets = FrozenDict({k: v[0] for k, v in sheets.items()})
        else:
            sheets = FrozenDict({k:v[0] for k,v in sheets.items() })
        missing_tables = {t for t in all_tables if t not in sheets}
        if missing_tables and print_missing_tables:
            print ("The following table names could not be found in the %s file.\n%s\n"%
                   (xls_file_path,"\n".join(missing_tables)))
        field_indicies, missing_fields, dup_fields = {}, {}, {}
        for table, sheet in sheets.items() :
            field_indicies[table], missing_fields[table], dup_fields[table] = \
                self._get_field_indicies(xls_file_path, table, sheet, row_offsets[table], headers_present)
        verify(not any(_ for _ in missing_fields.values()),
               "The following field names could not be found : \n" +
               "\n".join("%s : "%t + ",".join(bf) for t,bf in missing_fields.items() if bf))
        verify(not any(_ for _ in dup_fields.values()),
               "The following field names were duplicated : \n" +
               "\n".join("%s : "%t + ",".join(bf) for t,bf in dup_fields.items() if bf))
        return sheets, field_indicies, book.datemode
    def _create_generator_obj(self, xlsFilePath, table, row_offset, headers_present, treat_inf_as_infinity):
        tdf = self.tic_dat_factory
        ho = 1 if headers_present else 0
        def tableObj() :
            sheets, field_indicies, datemode = self._get_sheets_and_fields(xlsFilePath,
                                        (table,), {table:row_offset}, headers_present)
            if xlsFilePath.endswith('.xlsx'):
                if table in sheets :
                    sheet = sheets[table]
                    table_len = min(len(list(self.iter_cols(sheet))[field_indicies[table][field]])
                                   for field in tdf.data_fields[table])
                    row_list = list(self.iter_rows(sheet))
                    for x in (row_list[i] for i in range(table_len)[row_offset+ho:]):
                        yield self._sub_tuple(table, tdf.data_fields[table],
                                              field_indicies[table], treat_inf_as_infinity, datemode)(x)
            else:
                if table in sheets :
                    sheet = sheets[table]
                    table_len = min(len(sheet.col_values(field_indicies[table][field]))
                                   for field in tdf.data_fields[table])
                    for x in (sheet.row_values(i) for i in range(table_len)[row_offset+ho:]):
                        yield self._sub_tuple(table, tdf.data_fields[table],
                                              field_indicies[table], treat_inf_as_infinity, datemode)(x)
        return tableObj

    def _create_tic_dat_dict(self, xls_file_path, row_offsets, headers_present, treat_inf_as_infinity):
        tiai = treat_inf_as_infinity
        verify(utils.dictish(row_offsets) and
               set(row_offsets).issubset(self.tic_dat_factory.all_tables) and
               all(utils.numericish(x) and (x>=0) for x in row_offsets.values()),
               "row_offsets needs to map from table names to non negative row offset")
        row_offsets = dict({t:0 for t in self.tic_dat_factory.all_tables}, **row_offsets)
        tdf = self.tic_dat_factory
        rtn = {}
        sheets, field_indicies, dm = self._get_sheets_and_fields(xls_file_path,
                                    set(tdf.all_tables).difference(tdf.generator_tables),
                                    row_offsets, headers_present, print_missing_tables=True)
        ho = 1 if headers_present else 0
        if xls_file_path.endswith('.xlsx'):
            for tbl, sheet in sheets.items() :
                fields = tdf.primary_key_fields.get(tbl, ()) + tdf.data_fields.get(tbl, ())
                assert fields or tbl in self.tic_dat_factory.generic_tables
                indicies = field_indicies[tbl]
                table_len = min(len(list(self.iter_cols(sheet))[indicies[field]])
                                for field in (fields or indicies))
                if tdf.primary_key_fields.get(tbl, ()) :
                    row_list = list(self.iter_rows(sheet))
                    tableObj = {self._sub_tuple(tbl, tdf.primary_key_fields[tbl], indicies, tiai, dm)(x):
                                self._sub_tuple(tbl, tdf.data_fields.get(tbl, ()), indicies, tiai, dm)(x)
                                for x in (row_list[i] for i in
                                            range(table_len)[row_offsets[tbl]+ho:])}
                elif tbl in tdf.generic_tables:
                    tableObj = None # will be read via PanDatFactory
                else :
                    row_list = list(self.iter_rows(sheet))
                    tableObj = [self._sub_tuple(tbl, tdf.data_fields.get(tbl, ()), indicies, tiai, dm)(x)
                                for x in (row_list[i] for i in
                                            range(table_len)[row_offsets[tbl]+ho:])]
                if tableObj is not None:
                    rtn[tbl] = tableObj
        else:
            for tbl, sheet in sheets.items():
                fields = tdf.primary_key_fields.get(tbl, ()) + tdf.data_fields.get(tbl, ())
                assert fields or tbl in self.tic_dat_factory.generic_tables
                indicies = field_indicies[tbl]
                table_len = min(len(sheet.col_values(indicies[field]))
                                for field in (fields or indicies))
                if tdf.primary_key_fields.get(tbl, ()):
                    tableObj = {self._sub_tuple(tbl, tdf.primary_key_fields[tbl], indicies, tiai, dm)(x):
                                    self._sub_tuple(tbl, tdf.data_fields.get(tbl, ()), indicies, tiai, dm)(x)
                                for x in (sheet.row_values(i) for i in
                                          range(table_len)[row_offsets[tbl] + ho:])}
                elif tbl in tdf.generic_tables:
                    tableObj = None  # will be read via PanDatFactory
                else:
                    tableObj = [self._sub_tuple(tbl, tdf.data_fields.get(tbl, ()), indicies, tiai, dm)(x)
                                for x in (sheet.row_values(i) for i in
                                          range(table_len)[row_offsets[tbl] + ho:])]
                if tableObj is not None:
                    rtn[tbl] = tableObj
        for tbl in tdf.generator_tables :
            rtn[tbl] = self._create_generator_obj(xls_file_path, tbl, row_offsets[tbl],
                                                    headers_present, tiai)
        return rtn

    def find_duplicates(self, xls_file_path, row_offsets={}, headers_present = True):
        """
        Find the row counts for duplicated rows.

        :param xls_file_path: An Excel file containing sheets whose names match
                              the table names in the schema (non primary key tables ignored).

        :param row_offsets: (optional) A mapping from table names to initial
                            number of rows to skip (non primary key tables ignored)

        :param headers_present: Boolean. Does the first row of data contain the
                                column headers?

        caveats: Missing sheets resolve to an empty table, but missing primary fields
                 on matching sheets throw an Exception.
                 Sheet names are considered case insensitive.

        :return: A dictionary whose keys are the table names for the primary key tables.
                 Each value of the return dictionary is itself a dictionary.
                 The inner dictionary is keyed by the primary key values encountered
                 in the table, and the value is the count of records in the
                 Excel sheet with this primary key.
                 Row counts smaller than 2 are pruned off, as they aren't duplicates
        """
        self._verify_differentiable_sheet_names()
        verify(xlrd, "xlrd needs to be installed to use this subroutine")
        verify(utils.dictish(row_offsets) and
               set(row_offsets).issubset(self.tic_dat_factory.all_tables) and
               all(utils.numericish(x) and (x>=0) for x in row_offsets.values()),
               "row_offsets needs to map from table names to non negative row offset")
        row_offsets = dict({t:0 for t in self.tic_dat_factory.all_tables}, **row_offsets)
        tdf = self.tic_dat_factory
        pk_tables = tuple(t for t,_ in tdf.primary_key_fields.items() if _)
        rtn = {t:defaultdict(int) for t in pk_tables}
        sheets, fieldIndicies, dm = self._get_sheets_and_fields(xls_file_path, pk_tables,
                                        row_offsets, headers_present)
        ho = 1 if headers_present else 0
        if xls_file_path.endswith('.xlsx'):
            for table, sheet in sheets.items() :
                fields = tdf.primary_key_fields[table] + tdf.data_fields.get(table, ())
                indicies = fieldIndicies[table]
                table_len = min(len(list(self.iter_cols(sheet))[indicies[field]]) for field in fields)
                row_list = list(self.iter_rows(sheet))
                for x in (row_list[i] for i in range(table_len)[row_offsets[table]+ho:]) :
                    rtn[table][self._sub_tuple(table, tdf.primary_key_fields[table],
                                               indicies, treat_inf_as_infinity=True, datemode=dm)(x)] += 1
            for t in list(rtn.keys()):
                rtn[t] = {k:v for k,v in rtn[t].items() if v > 1}
                if not rtn[t]:
                    del(rtn[t])
        else:
            for table, sheet in sheets.items() :
                fields = tdf.primary_key_fields[table] + tdf.data_fields.get(table, ())
                indicies = fieldIndicies[table]
                table_len = min(len(sheet.col_values(indicies[field])) for field in fields)
                for x in (sheet.row_values(i) for i in range(table_len)[row_offsets[table]+ho:]) :
                    rtn[table][self._sub_tuple(table, tdf.primary_key_fields[table],
                                               indicies, treat_inf_as_infinity=True, datemode=dm)(x)] += 1
            for t in list(rtn.keys()):
                rtn[t] = {k:v for k,v in rtn[t].items() if v > 1}
                if not rtn[t]:
                    del(rtn[t])
        return rtn
    def _get_dv_dt(self, table, field):
        # reminder - data fields have a default default of zero, primary keys don't get a default default
        if (table, field) not in self._dv_dt:
            self._dv_dt[table, field] = (
                self.tic_dat_factory.default_values.get(table, {}).get(field, ["LIST", "NOT", "POSSIBLE"]),
                self.tic_dat_factory.data_types.get(table, {}).get(field)
            )
        return self._dv_dt[table, field]
    def _sub_tuple(self, table, fields, field_indicies, treat_inf_as_infinity, datemode) :
        assert set(fields).issubset(field_indicies)
        if self.tic_dat_factory.infinity_io_flag != "N/A" or \
            (table == "parameters" and self.tic_dat_factory.parameters):
            treat_inf_as_infinity = False
        def _read_cell(x, field):
            dv, dt = self._get_dv_dt(table, field)
            rtn = x[field_indicies[field]]
            if rtn == "" and ((dt and dt.nullable) or (not dt and dv is None)):
                return None
            if treat_inf_as_infinity and utils.stringish(rtn) and rtn.lower() in ["inf", "-inf"]:
                return float(rtn.lower())
            if utils.numericish(rtn) and utils.safe_apply(int)(rtn) == rtn and dt and dt.must_be_int:
                rtn = int(rtn)
            if rtn == "":
                try_rtn = self.tic_dat_factory._general_read_cell(table, field, None) # None as infinity flagging
                if utils.numericish(try_rtn):
                    return try_rtn
            if utils.numericish(rtn) and dt and dt.datetime:
                rtn = utils.safe_apply(lambda : xlrd.xldate_as_tuple(rtn, datemode))()
                if rtn is not None:
                    f = datetime.datetime
                    if utils.pd:
                        f = utils.pd.Timestamp
                    return f(year=rtn[0], month=rtn[1], day=rtn[2], hour=rtn[3], minute=rtn[4], second=rtn[5])
            return self.tic_dat_factory._general_read_cell(table, field, rtn)
        def rtn(x) :
            if len(fields) == 1 :
                return _read_cell(x, fields[0])
            return tuple(_read_cell(x, field) for field in fields)
        return rtn

    def iter_rows(self, ws):
        for row in ws.iter_rows():
            yield  [cell.value for cell in row]

    def iter_cols(self, ws):
        for col in ws.iter_cols():
            yield [cell.value for cell in col]

    def _get_field_indicies(self, xls_file_path, table, sheet, row_offset, headers_present) :
        fields = self.tic_dat_factory.primary_key_fields.get(table, ()) + \
                 self.tic_dat_factory.data_fields.get(table, ())
        if not headers_present:
            if xls_file_path.endswith('.xlsx'):
                row_len = len(list(self.iter_rows(sheet))) if sheet.max_row > 0 else len(fields)
                return ({f: i for i, f in enumerate(fields) if i < row_len},
                        [f for i, f in enumerate(fields) if i >= row_len], [])
            else:
                row_len = len(sheet.row_values(row_offset)) if sheet.nrows > 0  else len(fields)
                return ({f : i for i,f in enumerate(fields) if i < row_len},
                        [f for i,f in enumerate(fields) if i >= row_len], [])
        if xls_file_path.endswith('.xlsx'):
            if sheet.max_row - row_offset <= 0 :
                return {}, fields, []
        else:
            if sheet.nrows - row_offset <= 0 :
                return {}, fields, []
        if table in self.tic_dat_factory.generic_tables:
            temp_rtn = defaultdict(list)
            if xls_file_path.endswith('.xlsx'):
                for ind, val in enumerate(list(self.iter_rows(sheet))[row_offset]):
                    temp_rtn[val].append(ind)
            else:
                for ind, val in enumerate(sheet.row_values(row_offset)):
                    temp_rtn[val].append(ind)
        else:
            temp_rtn =  {field:list() for field in fields}
            if xls_file_path.endswith('.xlsx'):
                for field, (ind, val) in product(fields, enumerate(list(self.iter_rows(sheet))[row_offset])):
                    if field == val or (all(map(utils.stringish, (field, val))) and
                                        field.lower() == val.lower()):
                        temp_rtn[field].append(ind)
            else:
                for field, (ind, val) in product(fields, enumerate(sheet.row_values(row_offset))) :
                    if field == val or (all(map(utils.stringish, (field, val))) and
                                        field.lower() == val.lower()):
                        temp_rtn[field].append(ind)
        return ({field : inds[0] for field, inds in temp_rtn.items() if len(inds)==1},
                [field for field, inds in temp_rtn.items() if len(inds) == 0],
                [field for field, inds in temp_rtn.items() if len(inds) > 1])


    def write_file(self, tic_dat, file_path, allow_overwrite = False, case_space_sheet_names = False):
        """
        write the ticDat data to an excel file

        :param tic_dat: the data object to write (typically a TicDat)

        :param file_path: The file path of the excel file to create
                          Needs to end in either ".xls" or ".xlsx"
                          The latter is capable of writing out larger tables,
                          but the former handles infinity seamlessly.
                          If ".xlsx", then be advised that +/- float("inf") will be replaced
                          with "inf"/"-inf", unless infinity_io_flag is being applied.

        :param allow_overwrite: boolean - are we allowed to overwrite an
                                existing file?
              case_space_sheet_names: boolean - make best guesses how to add spaces and upper case
                                      characters to sheet names

        :return:

        caveats: None may be written out as an empty string. This reflects the behavior of xlwt.
        """
        self._verify_differentiable_sheet_names()
        verify(utils.stringish(file_path) and
               (file_path.endswith(".xls") or file_path.endswith(".xlsx")),
               "file_path argument needs to end in .xls or .xlsx")
        msg = []
        if not self.tic_dat_factory.good_tic_dat_object(tic_dat, lambda m : msg.append(m)) :
            raise TicDatError("Not a valid ticDat object for this schema : " + " : ".join(msg))
        verify(not os.path.isdir(file_path), "A directory is not a valid xls file path")
        verify(allow_overwrite or not os.path.exists(file_path),
               "The %s path exists and overwrite is not allowed"%file_path)
        if self.tic_dat_factory.generic_tables:
            dat, tdf = utils.create_generic_free(tic_dat, self.tic_dat_factory)
            return tdf.xls.write_file(dat, file_path, allow_overwrite, case_space_sheet_names)
        case_space_sheet_names = case_space_sheet_names and \
                                 len(set(self.tic_dat_factory.all_tables)) == \
                                 len(set(map(case_space_to_pretty, self.tic_dat_factory.all_tables)))
        tbl_name_mapping = {t:case_space_to_pretty(t) if case_space_sheet_names else t
                            for t in self.tic_dat_factory.all_tables}
        if file_path.endswith(".xls"):
            self._xls_write(tic_dat, file_path, tbl_name_mapping)
        else:
            self._xlsx_write(tic_dat, file_path, tbl_name_mapping)
    def _xls_write(self, tic_dat, file_path, tbl_name_mapping):
        verify(xlwt, "Can't write .xls files because xlwt package isn't installed.")
        tdf = self.tic_dat_factory
        def clean_for_write(t, f, x):
            if isinstance(x, datetime.datetime):
                return str(x)
            return self.tic_dat_factory._infinity_flag_write_cell(t, f, x)
        book = xlwt.Workbook()
        for t in  sorted(sorted(tdf.all_tables),
                         key=lambda x: len(tdf.primary_key_fields.get(x, ()))) :
            all_flds = self.tic_dat_factory.primary_key_fields.get(t, ()) + self.tic_dat_factory.data_fields.get(t, ())
            sheet = book.add_sheet(tbl_name_mapping[t][:_longest_sheet])
            for i,f in enumerate(tdf.primary_key_fields.get(t,()) + tdf.data_fields.get(t, ())) :
                sheet.write(0, i, f)
            _t = getattr(tic_dat, t)
            if utils.dictish(_t) :
                for row_ind, (p_key, data) in enumerate(_t.items()) :
                    for field_ind, cell in enumerate( (p_key if containerish(p_key) else (p_key,)) +
                                        tuple(data[_f] for _f in tdf.data_fields.get(t, ()))):
                        sheet.write(row_ind+1, field_ind, clean_for_write(t, all_flds[field_ind], cell))
            else :
                for row_ind, data in enumerate(_t if containerish(_t) else _t()) :
                    for field_ind, cell in enumerate(tuple(data[_f] for _f in tdf.data_fields[t])) :
                        sheet.write(row_ind+1, field_ind, clean_for_write(t, all_flds[field_ind], cell))
        if os.path.exists(file_path):
            os.remove(file_path)
        book.save(file_path)
    def _xlsx_write(self, tic_dat, file_path, tbl_name_mapping):
        verify(xlsx, "Can't write .xlsx files because xlsxwriter package isn't installed.")
        tdf = self.tic_dat_factory
        if os.path.exists(file_path):
            os.remove(file_path)
        book = xlsx.Workbook(file_path)
        def clean_for_write(t, f, x):
            if self.tic_dat_factory.infinity_io_flag != "N/A" or \
               (t == "parameters" and self.tic_dat_factory.parameters):
                return self.tic_dat_factory._infinity_flag_write_cell(t, f, x)
            if x in [float("inf"), -float("inf")] or isinstance(x, datetime.datetime):
                return str(x)
            return x
        for t in sorted(sorted(tdf.all_tables),
                         key=lambda x: len(tdf.primary_key_fields.get(x, ()))) :
            all_flds = self.tic_dat_factory.primary_key_fields.get(t, ()) + self.tic_dat_factory.data_fields.get(t, ())
            sheet = book.add_worksheet(tbl_name_mapping[t][:_longest_sheet])
            for i,f in enumerate(tdf.primary_key_fields.get(t,()) + tdf.data_fields.get(t, ())) :
                sheet.write(0, i, f)
            _t = getattr(tic_dat, t)
            if utils.dictish(_t) :
                for row_ind, (p_key, data) in enumerate(_t.items()) :
                    for field_ind, cell in enumerate( (p_key if containerish(p_key) else (p_key,)) +
                                        tuple(data[_f] for _f in tdf.data_fields.get(t, ()))):
                        sheet.write(row_ind+1, field_ind, clean_for_write(t, all_flds[field_ind], cell))
            else :
                for row_ind, data in enumerate(_t if containerish(_t) else _t()) :
                    for field_ind, cell in enumerate(tuple(data[_f] for _f in tdf.data_fields[t])) :
                        sheet.write(row_ind+1, field_ind, clean_for_write(t, all_flds[field_ind], cell))
        book.close()
